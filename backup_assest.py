from pathlib import Path
from openpyxl import load_workbook
from collections import defaultdict
import csv
import re
import shutil
import sys


# ============================================================
# EMKP FINAL ASSET CLEANUP + IN-PLACE RENAME
# ============================================================

SOURCE_ROOT = Path(r"E:\8-14_updated files")
EXCEL_FILE = Path(r"E:\16_Guide to the Dataset.xlsx")

PREFIX = "2024G07"


# ============================================================
# SAFETY
# ============================================================

# True  = preview only; NOTHING is deleted or renamed
# False = permanently delete excluded files/folder and rename
#
# KEEP TRUE FOR THE FIRST RUN.
DRY_RUN = False


# ============================================================
# FELIPE / BRITISH MUSEUM REQUESTS
# ============================================================

DELETE_EXTENSIONS = {
    ".dng",
    ".jpg",
    ".jpeg",
    ".arw",
}

DELETE_FOLDER_NAMES = {
    "15_EXTRAS",
}

# TXT files are intentionally retained.


# ============================================================
# AUDIT / ROLLBACK FILES
# ============================================================

WORK_ROOT = Path(r"E:\EMKP_work")

PLAN_CSV = (
    WORK_ROOT
    / "EMKP_final_plan.csv"
)

ROLLBACK_CSV = (
    WORK_ROOT
    / "EMKP_final_rename_rollback.csv"
)


# ============================================================
# HELPERS
# ============================================================

def clean_text(value):

    if value is None:
        return ""

    return str(value).strip()


def normalize_folder_name(name):

    return (
        " ".join(
            name.strip().split()
        )
        .casefold()
    )


def natural_text_parts(text):

    parts = re.split(
        r"(\d+)",
        text.casefold(),
    )

    return tuple(
        (0, int(part)) if part.isdigit() else (1, part)
        for part in parts
        if part
    )


def natural_sort_key(path, root_folder):

    relative_path = path.relative_to(root_folder)
    key = []

    for index, part in enumerate(relative_path.parts):

        if index == len(relative_path.parts) - 1:

            file_part = Path(part)

            key.append((
                natural_text_parts(file_part.stem),
                file_part.suffix.casefold(),
            ))

        else:

            key.append((
                natural_text_parts(part),
                "",
            ))

    return tuple(key)


def bytes_to_gb(value):

    return value / (1024 ** 3)


def normalize_final_extension(path):

    """
    Keep the original file format.

    Only normalize extension spelling/case.

    Examples:
        .TIF -> .tif
        .MP4 -> .mp4
        .PDF -> .pdf

    Also fixes malformed MP4 suffixes found
    during the original asset inventory.
    """

    extension = path.suffix.lower()

    if extension in {
        ".mp4_",
        ". mp4",
    }:
        return ".mp4"

    return extension


# ============================================================
# READ EXCEL MAPPING
# ============================================================

def read_excel_mapping(excel_file):

    workbook = load_workbook(
        excel_file,
        read_only=True,
        data_only=True,
    )

    worksheet = workbook.active

    header_row = None
    item_col = None
    folder_col = None


    for row_number, row in enumerate(
        worksheet.iter_rows(
            values_only=True
        ),
        start=1,
    ):

        values = [
            clean_text(value)
            for value in row
        ]


        for column_number, value in enumerate(
            values
        ):

            if value.casefold() == "item id":

                item_col = column_number


            elif (
                value.casefold()
                == "local file identifier"
            ):

                folder_col = column_number


        if (
            item_col is not None
            and folder_col is not None
        ):

            header_row = row_number
            break


    if header_row is None:

        raise ValueError(
            'Could not find "Item ID" and '
            '"Local File Identifier" columns.'
        )


    mapping = []


    for row in worksheet.iter_rows(
        min_row=header_row + 1,
        values_only=True,
    ):

        item_id = clean_text(
            row[item_col]
            if item_col < len(row)
            else None
        )


        folder_name = clean_text(
            row[folder_col]
            if folder_col < len(row)
            else None
        )


        if item_id and folder_name:

            mapping.append(
                (
                    folder_name,
                    item_id,
                )
            )


    return mapping


# ============================================================
# CREATE FOLDER INDEX
# ============================================================

def create_folder_index(source_root):

    index = {}


    for directory in source_root.rglob("*"):

        if directory.is_dir():

            key = normalize_folder_name(
                directory.name
            )

            index.setdefault(
                key,
                [],
            ).append(
                directory
            )


    root_key = normalize_folder_name(
        source_root.name
    )

    index.setdefault(
        root_key,
        [],
    ).append(
        source_root
    )


    return index


# ============================================================
# GET FILES FOR EACH MAPPED FOLDER
# ============================================================

def get_files_for_folder(
    folder,
    mapped_folders,
):

    files = []


    for path in folder.rglob("*"):

        if not path.is_file():
            continue


        parent = path.parent

        belongs_to_child_mapping = False


        while parent != folder:

            if parent in mapped_folders:

                belongs_to_child_mapping = True
                break

            parent = parent.parent


        if not belongs_to_child_mapping:

            files.append(
                path
            )


    files.sort(
        key=lambda path:
            natural_sort_key(
                path,
                folder,
            )
    )


    return files


# ============================================================
# WRITE PLAN CSV
# ============================================================

def write_plan_csv(records):

    WORK_ROOT.mkdir(
        parents=True,
        exist_ok=True,
    )


    with PLAN_CSV.open(
        "w",
        newline="",
        encoding="utf-8-sig",
    ) as file:

        writer = csv.writer(
            file
        )


        writer.writerow([
            "Sequence",
            "Item ID",
            "Folder",
            "Action",
            "Original Path",
            "Final Path",
            "Reason",
            "Size GB",
        ])


        for record in records:

            writer.writerow([
                record["sequence"],
                record["item_id"],
                record["folder_name"],
                record["action"],
                str(
                    record["source"]
                ),
                str(
                    record["target"]
                )
                if record["target"] is not None
                else "",
                record["reason"],
                round(
                    bytes_to_gb(
                        record["size_bytes"]
                    ),
                    6,
                ),
            ])


# ============================================================
# WRITE RENAME ROLLBACK CSV
# ============================================================

def write_rollback_csv(
    rename_operations
):

    WORK_ROOT.mkdir(
        parents=True,
        exist_ok=True,
    )


    with ROLLBACK_CSV.open(
        "w",
        newline="",
        encoding="utf-8-sig",
    ) as file:

        writer = csv.writer(
            file
        )


        writer.writerow([
            "Sequence",
            "Item ID",
            "Original Path",
            "Renamed Path",
        ])


        for operation in rename_operations:

            writer.writerow([
                operation[
                    "sequence"
                ],
                operation[
                    "item_id"
                ],
                str(
                    operation[
                        "source"
                    ]
                ),
                str(
                    operation[
                        "target"
                    ]
                ),
            ])


# ============================================================
# MAIN
# ============================================================

def main():

    print()

    print("=" * 84)

    print(
        "EMKP FINAL ASSET CLEANUP "
        "+ IN-PLACE RENAME"
    )

    print("=" * 84)


    print(
        f"Source:                "
        f"{SOURCE_ROOT}"
    )

    print(
        f"Excel:                 "
        f"{EXCEL_FILE}"
    )

    print(
        f"Prefix:                "
        f"{PREFIX}"
    )

    print(
        f"Dry run:               "
        f"{DRY_RUN}"
    )

    print(
        f"Work folder:           "
        f"{WORK_ROOT}"
    )

    print(
        f"Plan CSV:              "
        f"{PLAN_CSV}"
    )

    print(
        f"Rename rollback CSV:   "
        f"{ROLLBACK_CSV}"
    )

    print()


    # ========================================================
    # BASIC CHECKS
    # ========================================================

    if not SOURCE_ROOT.exists():

        print(
            "ERROR: Source directory "
            "does not exist:"
        )

        print(
            SOURCE_ROOT
        )

        sys.exit(1)


    if not EXCEL_FILE.exists():

        print(
            "ERROR: Excel file "
            "does not exist:"
        )

        print(
            EXCEL_FILE
        )

        sys.exit(1)


    # ========================================================
    # TEST WRITE ACCESS
    # ========================================================

    try:

        WORK_ROOT.mkdir(
            parents=True,
            exist_ok=True,
        )

    except Exception as error:

        print()

        print(
            "ERROR: Cannot create "
            "the work directory:"
        )

        print(
            WORK_ROOT
        )

        print()

        print(error)

        sys.exit(1)


    # ========================================================
    # READ EXCEL MAPPING
    # ========================================================

    mapping = read_excel_mapping(
        EXCEL_FILE
    )


    print(
        f"Found {len(mapping)} "
        f"Excel folder mappings."
    )


    # ========================================================
    # RESOLVE FOLDERS
    # ========================================================

    folder_index = create_folder_index(
        SOURCE_ROOT
    )


    resolved = []

    errors = []


    for (
        folder_name,
        item_id,
    ) in mapping:


        key = normalize_folder_name(
            folder_name
        )


        matches = folder_index.get(
            key,
            [],
        )


        if len(matches) == 0:

            errors.append(
                f"NOT FOUND: "
                f"{folder_name} -> "
                f"{item_id}"
            )


        elif len(matches) > 1:

            message = (
                f"AMBIGUOUS FOLDER: "
                f"{folder_name} -> "
                f"{item_id}\n"
            )


            for match in matches:

                message += (
                    f"    {match}\n"
                )


            errors.append(
                message
            )


        else:

            resolved.append(
                (
                    matches[0],
                    item_id,
                    folder_name,
                )
            )


    # ========================================================
    # STOP IF MAPPING HAS PROBLEMS
    # ========================================================

    if errors:

        print()

        print("=" * 84)

        print(
            "FOLDER MAPPING ERRORS"
        )

        print("=" * 84)


        for error in errors:

            print(error)


        print()

        print(
            "Nothing was changed."
        )


        sys.exit(1)


    mapped_folders = {

        folder

        for (
            folder,
            _,
            _,
        )

        in resolved
    }


    # ========================================================
    # BUILD FINAL PLAN
    # ========================================================

    all_records = []

    rename_operations = []

    excluded_file_operations = []

    folders_to_delete = set()


    # --------------------------------------------------------
    # IMPORTANT
    #
    # Sequence numbering remains based on the original
    # 2,469-asset verification list.
    #
    # Deleted files leave intentional numbering gaps.
    #
    # Retained assets are NOT renumbered.
    # --------------------------------------------------------

    sequence = 1


    for (
        folder,
        item_id,
        folder_name,
    ) in resolved:


        files = get_files_for_folder(
            folder,
            mapped_folders,
        )


        delete_entire_folder = (
            folder_name
            in DELETE_FOLDER_NAMES
        )


        if delete_entire_folder:

            folders_to_delete.add(
                folder
            )


        for source in files:


            try:

                size_bytes = (
                    source.stat().st_size
                )

            except OSError:

                size_bytes = 0


            extension = (
                source.suffix.lower()
            )


            # =================================================
            # DELETE ENTIRE 15_EXTRAS FOLDER
            # =================================================

            if delete_entire_folder:


                record = {

                    "sequence":
                        sequence,

                    "item_id":
                        item_id,

                    "folder_name":
                        folder_name,

                    "action":
                        "DELETE",

                    "source":
                        source,

                    "target":
                        None,

                    "reason":
                        (
                            "Entire 15_EXTRAS folder "
                            "deleted per Felipe"
                        ),

                    "size_bytes":
                        size_bytes,
                }


                all_records.append(
                    record
                )


            # =================================================
            # DELETE UNSUPPORTED FILE TYPES
            # =================================================

            elif (
                extension
                in DELETE_EXTENSIONS
            ):


                record = {

                    "sequence":
                        sequence,

                    "item_id":
                        item_id,

                    "folder_name":
                        folder_name,

                    "action":
                        "DELETE",

                    "source":
                        source,

                    "target":
                        None,

                    "reason":
                        (
                            "Unsupported British Museum "
                            f"extension: {extension}"
                        ),

                    "size_bytes":
                        size_bytes,
                }


                all_records.append(
                    record
                )


                excluded_file_operations.append(
                    record
                )


            # =================================================
            # KEEP AND RENAME
            # =================================================

            else:


                final_extension = (
                    normalize_final_extension(
                        source
                    )
                )


                final_filename = (

                    f"{PREFIX}-"

                    f"{item_id}-"

                    f"{sequence:04d}"

                    f"{final_extension}"
                )


                target = source.with_name(
                    final_filename
                )


                record = {

                    "sequence":
                        sequence,

                    "item_id":
                        item_id,

                    "folder_name":
                        folder_name,

                    "action":
                        "RENAME",

                    "source":
                        source,

                    "target":
                        target,

                    "reason":
                        "",

                    "size_bytes":
                        size_bytes,
                }


                all_records.append(
                    record
                )


                rename_operations.append(
                    record
                )


            sequence += 1


    # ========================================================
    # SUMMARY
    # ========================================================

    total_assets = len(
        all_records
    )


    deleted_assets = sum(

        1

        for record
        in all_records

        if (
            record["action"]
            == "DELETE"
        )
    )


    retained_assets = len(
        rename_operations
    )


    total_bytes = sum(

        record["size_bytes"]

        for record
        in all_records
    )


    deleted_bytes = sum(

        record["size_bytes"]

        for record
        in all_records

        if (
            record["action"]
            == "DELETE"
        )
    )


    retained_bytes = sum(

        record["size_bytes"]

        for record
        in rename_operations
    )


    # ========================================================
    # RETAINED TYPE SUMMARY
    # ========================================================

    type_statistics = defaultdict(
        lambda: {
            "count": 0,
            "bytes": 0,
        }
    )


    for record in rename_operations:


        extension = (
            normalize_final_extension(
                record["source"]
            )
        )


        asset_type = (
            extension
            .lstrip(".")
            .upper()
        )


        if not asset_type:

            asset_type = (
                "NO_EXTENSION"
            )


        type_statistics[
            asset_type
        ]["count"] += 1


        type_statistics[
            asset_type
        ]["bytes"] += (
            record["size_bytes"]
        )


    # ========================================================
    # PRINT SUMMARY
    # ========================================================

    print()

    print("=" * 84)

    print(
        "FINAL PLAN SUMMARY"
    )

    print("=" * 84)


    print(
        f"Original assets:       "
        f"{total_assets:,}"
    )


    print(
        f"Assets to delete:      "
        f"{deleted_assets:,}"
    )


    print(
        f"Retained / renamed:    "
        f"{retained_assets:,}"
    )


    print(
        f"Original size:         "
        f"{bytes_to_gb(total_bytes):,.2f} GB"
    )


    print(
        f"Delete size:           "
        f"{bytes_to_gb(deleted_bytes):,.2f} GB"
    )


    print(
        f"Retained size:         "
        f"{bytes_to_gb(retained_bytes):,.2f} GB"
    )


    print()


    print(
        "RETAINED ASSET TYPES"
    )


    print("-" * 60)


    print(
        f"{'Type':<12}"
        f"{'Assets':>12}"
        f"{'Size (GB)':>16}"
    )


    print("-" * 60)


    for (
        asset_type,
        statistics,
    ) in sorted(

        type_statistics.items(),

        key=lambda item:
            item[1]["bytes"],

        reverse=True,
    ):


        print(

            f"{asset_type:<12}"

            f"{statistics['count']:>12,}"

            f"{bytes_to_gb(statistics['bytes']):>16,.2f}"
        )


    print()


    print(
        "DELETIONS"
    )


    print("-" * 60)


    print(
        "1. Entire folder: 15_EXTRAS"
    )


    print(
        "2. All .DNG files"
    )


    print(
        "3. All .JPG/.JPEG files"
    )


    print(
        "4. All .ARW files"
    )


    print(
        "5. .TXT files are retained"
    )


    print()


    # ========================================================
    # EXPECTED TOTAL VALIDATION
    # ========================================================

    if total_assets != 2469:

        print(
            "ERROR:"
        )

        print(
            "Expected 2,469 original assets "
            f"but found {total_assets:,}."
        )

        print(
            "Nothing was changed."
        )

        sys.exit(1)


    if deleted_assets != 591:

        print(
            "ERROR:"
        )

        print(
            "Expected 591 deleted assets "
            f"but found {deleted_assets:,}."
        )

        print(
            "Nothing was changed."
        )

        sys.exit(1)


    if retained_assets != 1878:

        print(
            "ERROR:"
        )

        print(
            "Expected 1,878 retained assets "
            f"but found {retained_assets:,}."
        )

        print(
            "Nothing was changed."
        )

        sys.exit(1)


    # ========================================================
    # DUPLICATE TARGET CHECK
    # ========================================================

    source_paths = {

        operation["source"]

        for operation
        in rename_operations
    }


    destination_paths = [

        operation["target"]

        for operation
        in rename_operations
    ]


    if (
        len(destination_paths)
        != len(
            set(destination_paths)
        )
    ):

        print(
            "ERROR: Duplicate destination "
            "filenames were generated."
        )

        print(
            "Nothing was changed."
        )

        sys.exit(1)


    # ========================================================
    # EXISTING TARGET COLLISION CHECK
    # ========================================================

    collisions = []


    for operation in rename_operations:


        destination = (
            operation["target"]
        )


        if (
            destination.exists()
            and destination
            not in source_paths
        ):

            collisions.append(
                destination
            )


    if collisions:

        print()

        print(
            "ERROR: Some proposed final "
            "filenames already exist:"
        )


        for collision in collisions[:30]:

            print(
                collision
            )


        print()

        print(
            "Nothing was changed."
        )

        sys.exit(1)


    # ========================================================
    # WRITE AUDIT PLAN
    # ========================================================

    try:

        write_plan_csv(
            all_records
        )


    except PermissionError:

        print()

        print(
            "ERROR: Could not write:"
        )

        print(
            PLAN_CSV
        )

        print()

        print(
            "Make sure EMKP_final_plan.csv "
            "is not open in Excel."
        )

        print()

        print(
            "Nothing was changed."
        )

        sys.exit(1)


    print(
        "Plan written:"
    )

    print(
        f"  {PLAN_CSV}"
    )


    # ========================================================
    # DRY RUN STOPS HERE
    # ========================================================

    if DRY_RUN:


        print()

        print("=" * 84)


        print(
            "DRY RUN COMPLETE — "
            "NOTHING CHANGED"
        )


        print("=" * 84)


        print()

        print(
            "Expected:"
        )


        print(
            "  Original assets:    2,469"
        )


        print(
            "  Delete assets:        591"
        )


        print(
            "  Retained assets:     1,878"
        )


        print(
            "  Retained size:      ~187.21 GB"
        )


        print()

        print(
            "Review:"
        )


        print(
            f"  {PLAN_CSV}"
        )


        print()

        print(
            "When ready for the real run, "
            "change:"
        )


        print()

        print(
            "    DRY_RUN = True"
        )


        print()

        print(
            "to:"
        )


        print()

        print(
            "    DRY_RUN = False"
        )


        print()

        return


    # ========================================================
    # REAL RUN STARTS HERE
    # ========================================================

    print()

    print("=" * 84)

    print(
        "STARTING PERMANENT CLEANUP "
        "AND RENAME"
    )

    print("=" * 84)


    # ========================================================
    # WRITE ROLLBACK BEFORE RENAMING
    # ========================================================

    write_rollback_csv(
        rename_operations
    )


    print()

    print(
        "Rename rollback file created:"
    )


    print(
        f"  {ROLLBACK_CSV}"
    )


    print()

    print(
        "IMPORTANT:"
    )

    print(
        "The rollback CSV can restore renamed "
        "filenames, but it cannot restore "
        "permanently deleted files."
    )


    # ========================================================
    # STEP 1:
    # DELETE 15_EXTRAS
    # ========================================================

    print()

    print(
        "Deleting 15_EXTRAS..."
    )


    for folder in sorted(

        folders_to_delete,

        key=lambda path:
            len(path.parts),

        reverse=True,
    ):


        if not folder.exists():
            continue


        print(
            f"Deleting folder:"
        )

        print(
            f"  {folder}"
        )


        shutil.rmtree(
            folder
        )


    # ========================================================
    # STEP 2:
    # DELETE DNG / JPG / JPEG / ARW
    # ========================================================

    print()

    print(
        "Deleting unsupported "
        "DNG/JPG/JPEG/ARW files..."
    )


    deleted_file_count = 0


    for operation in (
        excluded_file_operations
    ):


        source = (
            operation["source"]
        )


        # Files inside 15_EXTRAS were already
        # deleted with the whole folder.
        if not source.exists():

            continue


        source.unlink()


        deleted_file_count += 1


        if (
            deleted_file_count % 25
            == 0
        ):

            print(

                f"Unsupported files deleted: "
                f"{deleted_file_count}"
            )


    print(

        f"Unsupported files deleted "
        f"outside 15_EXTRAS: "
        f"{deleted_file_count}"
    )


    # ========================================================
    # STEP 3:
    # TWO-PHASE RENAME
    # ========================================================

    print()

    print(
        "Renaming retained assets..."
    )


    temporary_operations = []


    try:


        # ----------------------------------------------------
        # PHASE 1:
        # ORIGINAL -> TEMPORARY
        # ----------------------------------------------------

        for (
            index,
            operation,
        ) in enumerate(

            rename_operations,

            start=1,
        ):


            source = (
                operation["source"]
            )


            if not source.exists():

                raise FileNotFoundError(
                    f"Expected retained asset "
                    f"not found:\n"
                    f"{source}"
                )


            temporary_name = (

                f".__emkp_tmp__"

                f"{index:06d}__"

                f"{source.name}"
            )


            temporary_path = (
                source.with_name(
                    temporary_name
                )
            )


            if temporary_path.exists():

                raise FileExistsError(
                    f"Temporary path already "
                    f"exists:\n"
                    f"{temporary_path}"
                )


            source.rename(
                temporary_path
            )


            temporary_operations.append(
                (
                    temporary_path,
                    operation["target"],
                )
            )


            if (
                index % 100 == 0
                or index
                == len(
                    rename_operations
                )
            ):

                print(

                    f"Temporary phase: "

                    f"{index:,}/"

                    f"{len(rename_operations):,}"
                )


        # ----------------------------------------------------
        # PHASE 2:
        # TEMPORARY -> FINAL
        # ----------------------------------------------------

        for (
            index,
            (
                temporary_path,
                final_path,
            ),
        ) in enumerate(

            temporary_operations,

            start=1,
        ):


            if final_path.exists():

                raise FileExistsError(
                    f"Final target unexpectedly "
                    f"exists:\n"
                    f"{final_path}"
                )


            temporary_path.rename(
                final_path
            )


            if (
                index % 100 == 0
                or index
                == len(
                    temporary_operations
                )
            ):

                print(

                    f"Final phase: "

                    f"{index:,}/"

                    f"{len(temporary_operations):,}"
                )


    except Exception as error:


        print()

        print("=" * 84)

        print(
            "ERROR DURING RENAME"
        )

        print("=" * 84)


        print(error)


        print()

        print(
            "DO NOT delete these files:"
        )


        print(
            f"  {PLAN_CSV}"
        )


        print(
            f"  {ROLLBACK_CSV}"
        )


        print()

        print(
            "Some files may have temporary "
            "names if the script stopped "
            "during the rename process."
        )


        sys.exit(2)


    # ========================================================
    # SUCCESS
    # ========================================================

    print()

    print("=" * 84)

    print(
        "SUCCESS"
    )

    print("=" * 84)


    print(
        f"Assets retained / renamed: "
        f"{len(rename_operations):,}"
    )


    print(
        f"Assets permanently deleted: "
        f"{deleted_assets:,}"
    )


    print()


    print(
        "Final dataset:"
    )


    print(
        f"  {SOURCE_ROOT}"
    )


    print()


    print(
        "Audit plan:"
    )


    print(
        f"  {PLAN_CSV}"
    )


    print()


    print(
        "Rename rollback:"
    )


    print(
        f"  {ROLLBACK_CSV}"
    )


    print()


    print(
        "15_EXTRAS was permanently deleted."
    )


    print(
        "DNG, JPG/JPEG, and ARW files "
        "were permanently deleted."
    )


    print(
        "TXT files were retained."
    )


    print(
        "Remaining file contents "
        "were not converted."
    )


    print(
        "The original reviewed sequence "
        "numbers were preserved."
    )


# ============================================================
# START
# ============================================================

if __name__ == "__main__":

    main()
