from pathlib import Path
from openpyxl import load_workbook
import csv
import re
import sys

# ============================================================
# SETTINGS
# ============================================================

SOURCE_ROOT = Path(r"E:\8-14_updated files")
EXCEL_FILE = Path(r"E:\16_Guide to the Dataset.xlsx")

PREFIX = "2024G07"

# SAFETY:
# True  = preview only, no files renamed
# False = rename files in place
DRY_RUN = False

# Rollback plan is written BEFORE any real rename starts.
ROLLBACK_CSV = Path(r"E:\EMKP_work\rename_rollback-Aug18.csv")


# ============================================================
# HELPERS
# ============================================================

def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_folder_name(name):
    return " ".join(name.strip().split()).casefold()


def natural_text_parts(text):
    parts = re.split(r"(\d+)", text.casefold())

    return tuple(
        (0, int(part)) if part.isdigit() else (1, part)
        for part in parts
        if part
    )


def natural_sort_key(path, root_folder):
    relative_path = path.relative_to(root_folder)
    key = []

    for index, part in enumerate(relative_path.parts):
        if index == len(relative_path.parts) - 1:
            file_part = Path(part)
            key.append((
                natural_text_parts(file_part.stem),
                file_part.suffix.casefold(),
            ))
        else:
            key.append((
                natural_text_parts(part),
                "",
            ))

    return tuple(key)


def read_excel_mapping(excel_file):
    wb = load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.active

    header_row = None
    item_id_column = None
    folder_column = None

    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [clean_text(v) for v in row]

        for column_number, value in enumerate(values):
            if value.casefold() == "item id":
                item_id_column = column_number
            elif value.casefold() == "local file identifier":
                folder_column = column_number

        if item_id_column is not None and folder_column is not None:
            header_row = row_number
            break

    if header_row is None:
        raise ValueError(
            'Could not find "Item ID" and "Local File Identifier" columns.'
        )

    mapping = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        item_id = clean_text(
            row[item_id_column] if item_id_column < len(row) else None
        )
        folder_name = clean_text(
            row[folder_column] if folder_column < len(row) else None
        )

        if item_id and folder_name:
            mapping.append((folder_name, item_id))

    return mapping


def create_folder_index(source_root):
    index = {}

    for directory in source_root.rglob("*"):
        if directory.is_dir():
            key = normalize_folder_name(directory.name)
            index.setdefault(key, []).append(directory)

    key = normalize_folder_name(source_root.name)
    index.setdefault(key, []).append(source_root)

    return index


def get_files_for_folder(folder, mapped_folders):
    files = []

    for path in folder.rglob("*"):
        if not path.is_file():
            continue

        # Do not absorb files from another mapped child folder.
        belongs_to_child_mapping = False
        parent = path.parent

        while parent != folder:
            if parent in mapped_folders:
                belongs_to_child_mapping = True
                break
            parent = parent.parent

        if not belongs_to_child_mapping:
            files.append(path)

    files.sort(
        key=lambda p: natural_sort_key(p, folder)
    )

    return files


def get_extension(file_path):
    # Preserve existing extension, normalized to lowercase in new filename.
    return file_path.suffix.lower()


def write_rollback_csv(operations):
    """
    Write the rollback file BEFORE renaming begins.
    """
    with ROLLBACK_CSV.open(
        "w",
        newline="",
        encoding="utf-8-sig"
    ) as f:
        writer = csv.writer(f)
        writer.writerow([
            "Sequence",
            "Item ID",
            "Original Path",
            "Renamed Path",
            "Original Filename",
            "New Filename",
        ])

        for op in operations:
            writer.writerow([
                op["sequence"],
                op["item_id"],
                str(op["source"]),
                str(op["destination"]),
                op["source"].name,
                op["destination"].name,
            ])


# ============================================================
# MAIN
# ============================================================

def main():
    print()
    print("=" * 78)
    print("IN-PLACE DATASET RENAMING")
    print("=" * 78)
    print(f"Source:       {SOURCE_ROOT}")
    print(f"Excel:        {EXCEL_FILE}")
    print(f"Prefix:       {PREFIX}")
    print(f"Dry run:      {DRY_RUN}")
    print(f"Rollback CSV: {ROLLBACK_CSV}")
    print()

    if not SOURCE_ROOT.exists():
        print(f"ERROR: Source directory does not exist: {SOURCE_ROOT}")
        sys.exit(1)

    if not EXCEL_FILE.exists():
        print(f"ERROR: Excel file does not exist: {EXCEL_FILE}")
        sys.exit(1)

    mapping = read_excel_mapping(EXCEL_FILE)
    print(f"Found {len(mapping)} Excel folder mappings.")

    directory_index = create_folder_index(SOURCE_ROOT)

    resolved = []
    errors = []

    for folder_name, item_id in mapping:
        key = normalize_folder_name(folder_name)
        matches = directory_index.get(key, [])

        if len(matches) == 0:
            errors.append(f"NOT FOUND: {folder_name} -> {item_id}")
        elif len(matches) > 1:
            msg = f"AMBIGUOUS FOLDER: {folder_name} -> {item_id}\n"
            for match in matches:
                msg += f"    {match}\n"
            errors.append(msg)
        else:
            resolved.append((matches[0], item_id, folder_name))

    if errors:
        print()
        print("=" * 78)
        print("ERRORS FOUND")
        print("=" * 78)
        for error in errors:
            print(error)
        print("\nNo files were renamed.")
        sys.exit(1)

    mapped_folders = {folder for folder, _, _ in resolved}

    operations = []
    sequence = 1

    print()
    print("=" * 78)
    print("RENAME PLAN")
    print("=" * 78)

    for folder, item_id, excel_folder_name in resolved:
        files = get_files_for_folder(folder, mapped_folders)
        rel_folder = folder.relative_to(SOURCE_ROOT)

        print(f"\n{item_id} | {rel_folder}")
        print(f"Files found: {len(files)}")

        for source in files:
            extension = get_extension(source)

            new_name = (
                f"{PREFIX}-"
                f"{item_id}-"
                f"{sequence:04d}"
                f"{extension}"
            )

            destination = source.with_name(new_name)

            operations.append({
                "sequence": sequence,
                "item_id": item_id,
                "source": source,
                "destination": destination,
            })

            print(
                f"{sequence:04d}: "
                f"{source.name}  -->  {new_name}"
            )

            sequence += 1

    print()
    print("=" * 78)
    print(f"TOTAL FILES TO RENAME: {len(operations):,}")
    print("=" * 78)

    # --------------------------------------------------------
    # PRE-FLIGHT SAFETY CHECKS
    # --------------------------------------------------------

    source_paths = {op["source"] for op in operations}
    destination_paths = [op["destination"] for op in operations]

    if len(destination_paths) != len(set(destination_paths)):
        print("ERROR: Duplicate destination filenames were generated.")
        print("No files were renamed.")
        sys.exit(1)

    collisions = []

    for op in operations:
        dst = op["destination"]

        # It is okay if a destination is another source path that will also
        # be renamed, because we use temporary filenames during the real run.
        if dst.exists() and dst not in source_paths:
            collisions.append(dst)

    if collisions:
        print("\nERROR: Some target filenames already exist and are NOT part of this rename set.")
        for path in collisions[:25]:
            print(f"  {path}")
        print("\nNo files were renamed.")
        sys.exit(1)

    if DRY_RUN:
        print()
        print("DRY RUN COMPLETE.")
        print("No files were renamed.")
        print()
        print("If the plan is correct, change:")
        print("    DRY_RUN = True")
        print("to:")
        print("    DRY_RUN = False")
        print()
        return

    # --------------------------------------------------------
    # WRITE ROLLBACK PLAN FIRST
    # --------------------------------------------------------

    write_rollback_csv(operations)

    print()
    print(f"Rollback plan created: {ROLLBACK_CSV}")
    print()

    # --------------------------------------------------------
    # SAFE TWO-PHASE RENAME
    #
    # Phase 1: source -> unique temporary name
    # Phase 2: temporary -> final name
    #
    # This avoids collisions when old filenames happen to equal
    # another file's future filename.
    # --------------------------------------------------------

    temp_operations = []

    print("=" * 78)
    print("RENAMING FILES")
    print("=" * 78)

    try:
        # Phase 1
        for i, op in enumerate(operations, start=1):
            src = op["source"]

            temp_name = (
                f".__rename_tmp__{i:06d}__"
                f"{src.name}"
            )

            temp_path = src.with_name(temp_name)

            if temp_path.exists():
                raise FileExistsError(
                    f"Temporary path already exists: {temp_path}"
                )

            src.rename(temp_path)

            temp_operations.append({
                "temp": temp_path,
                "final": op["destination"],
                "original": src,
            })

            if i % 100 == 0 or i == len(operations):
                print(
                    f"Temporary phase: "
                    f"{i:,}/{len(operations):,}"
                )

        # Phase 2
        for i, op in enumerate(temp_operations, start=1):
            temp_path = op["temp"]
            final_path = op["final"]

            if final_path.exists():
                raise FileExistsError(
                    f"Final target unexpectedly exists: {final_path}"
                )

            temp_path.rename(final_path)

            if i % 100 == 0 or i == len(temp_operations):
                print(
                    f"Final phase: "
                    f"{i:,}/{len(temp_operations):,}"
                )

    except Exception as error:
        print()
        print("ERROR DURING RENAME:")
        print(error)
        print()
        print(
            "Some files may already have temporary or final names. "
            "Do NOT delete the rollback CSV."
        )
        print(
            "Use the rollback script provided with this rename script "
            "before making further changes."
        )
        sys.exit(2)

    print()
    print("=" * 78)
    print("SUCCESS")
    print("=" * 78)
    print(f"Renamed files: {len(operations):,}")
    print(f"Source directory: {SOURCE_ROOT}")
    print(f"Rollback CSV: {ROLLBACK_CSV}")
    print()
    print("No file contents were modified.")
    print("Folder names and nested folder structure were left unchanged.")


if __name__ == "__main__":
    main()
